import os
import sys
import subprocess
<<<<<<< Updated upstream
import calendar
import datetime
from datetime import date
=======
import datetime
>>>>>>> Stashed changes
import openpyxl
import pyodbc
from pathlib import Path
from fastapi import APIRouter, Depends, HTTPException

<<<<<<< Updated upstream

from app.core.database import get_db_connection
from app.schemas.mgmt_prod_report import MgmtProdReportRequest, ProdReportType2Request, OpenFileRequest
from app.constant.queris import (
    SaarthiMickyReportTCFBIW,
    LineStopRecordDaily,
    LineStopRecordMonthly,
=======
from app.core.database import get_db_connection
from app.schemas.mgmt_prod_report import ProdReportType2Request, OpenFileRequest
from app.constant.queris import (
    SaarthiMickyReportTCFBIW1,        # monthly / daily / YTD / FY aggregates
    SaarthiMickyWeekwiseMonthly,       # ISO week-by-week actuals for the month
    LineStopRecordDaily,
>>>>>>> Stashed changes
)

router = APIRouter()

BASE_DIR = Path(__file__).resolve().parents[3]
TEMPLATE_PATH = BASE_DIR / "ProductionReport_R3.xlsx"
<<<<<<< Updated upstream
# OUTPUT_PATH will be defined dynamically after parsing dates


@router.post("/prod-report-type2")
def generate_mgmt_production_report(
    payload: ProdReportType2Request,
    conn: pyodbc.Connection = Depends(get_db_connection)
):
    """
    Generates and populates the Management Production Report for Micky (M_TCF).
    Reads date parameters dynamically from request body, queries raw daily log data,
    performs Daily, Weekly, Monthly, MTD, YTD, and FY calculations, updates the Excel spreadsheet,
    and returns the updated report as a downloadable file.
    """
    # Helper to safely convert None to 0
=======

# ---------------------------------------------------------------------------
# SaarthiMickyReportTCFBIW1 -- actual result column index map (22 cols, 0-21)
# ---------------------------------------------------------------------------
#  [00]  DT
#  [01]  DAILY_TARGET
#  [02]  DAILY_ACTUAL
#  [03]  W1_TARGET    [04]  W1_ACTUAL
#  [05]  W2_TARGET    [06]  W2_ACTUAL
#  [07]  W3_TARGET    [08]  W3_ACTUAL
#  [09]  W4_TARGET    [10]  W4_ACTUAL
#  [11]  W5_TARGET    [12]  W5_ACTUAL   (WeekNo >= 5 merged)
#  [13]  MONTHLY_TARGET
#  [14]  MONTHLY_ACTUAL
#  [15]  TARGET_PREVIOUS_FINANCIAL_YEAR
#  [16]  ACTUAL_PREVIOUS_FINANCIAL_YEAR
#  [17]  MTD_TARGET
#  [18]  MTD_ACTUAL
#  [19]  YTD_TARGET
#  [20]  YTD_ACTUAL
#  [21]  DAYS_IN_MONTH
# ---------------------------------------------------------------------------
IDX_DT              = 0
IDX_DAILY_TARGET    = 1
IDX_DAILY_ACTUAL    = 2
IDX_MONTHLY_TARGET  = 13
IDX_MONTHLY_ACTUAL  = 14
IDX_PFY_ACTUAL      = 16   # Previous Financial Year Actual -> FY column (R)
IDX_YTD_ACTUAL      = 20   # Year-to-Date Actual            -> YTD column (P)
IDX_DAYS_IN_MONTH   = 21   # max valid index = 21

# SaarthiMickyWeekwiseMonthly returns 3 cols per row
IDX_WW_ISOWEEK  = 0   # ISO week number (int)
IDX_WW_TARGET   = 1   # WK_TARGET
IDX_WW_ACTUAL   = 2   # WK_ACTUAL

# Excel columns for the 5 week slots (H I J K L)
WEEK_EXCEL_COLS = [8, 9, 10, 11, 12]


def get_month_iso_weeks(report_date: datetime.date) -> list:
    """
    Return a sorted list of ISO 8601 week numbers that have at least one day
    in the same calendar month as report_date.

    Example: June 2026  ->  [23, 24, 25, 26, 27]
    A month can span 4, 5 or (rarely) 6 ISO weeks.
    """
    year, month = report_date.year, report_date.month
    seen = set()
    weeks = []
    d = datetime.date(year, month, 1)
    while d.month == month:
        wk = d.isocalendar()[1]
        if wk not in seen:
            seen.add(wk)
            weeks.append(wk)
        d += datetime.timedelta(days=1)
    return sorted(weeks)


def build_weekwise_dict(weekwise_rows) -> dict:
    """
    Convert weekwise query rows into a dict: {iso_week: (wk_target, wk_actual)}.
    Missing weeks get (0, 0) by default via .get().
    """
    return {int(r[IDX_WW_ISOWEEK]): (int(r[IDX_WW_TARGET] or 0), int(r[IDX_WW_ACTUAL] or 0))
            for r in weekwise_rows}


@router.post("/prod-report-type2")
def generate_prod_report_type2(
    payload: ProdReportType2Request,
    conn: pyodbc.Connection = Depends(get_db_connection),
):
    """
    Generates and populates the Production Report Type-2 (ProductionReport_R3.xlsx).

    Weekly columns (H-L) are dynamic: they show the actual ISO calendar week
    numbers for the report month (e.g. W23, W24, W25, W26, W27 for June).

    Section mapping:
        Vehicle (rows 10-13): Saarthi->S_TCF, Micky->M_TCF, Cargo/I-Puma->0
        BIW     (rows 19-21): Saarthi->S_BIW, Cargo/I-Puma->0
        Paint   (rows 27-30): all->0 (tables not yet live)
        Station Call (35-39): from LineStopRecordDaily

    Column layout per vehicle/BIW row:
        E(5)  = Monthly Plan    (MONTHLY_TARGET idx 13)
        F(6)  = Monthly Actual  (MONTHLY_ACTUAL idx 14)
        G(7)  = Monthly Var     (Actual - Plan)
        H(8)  = WXX Actual      (1st ISO week of month)
        I(9)  = WXX Actual      (2nd ISO week of month)
        J(10) = WXX Actual      (3rd ISO week of month)
        K(11) = WXX Actual      (4th ISO week of month)
        L(12) = WXX Actual      (5th ISO week of month, if exists)
        M(13) = Daily Plan      (DAILY_TARGET idx 1)
        N(14) = Daily Actual    (DAILY_ACTUAL idx 2)
        O(15) = Daily Var       (Actual - Plan)
        P(16) = YTD Actual      (YTD_ACTUAL idx 20)
        R(18) = FY Actual       (ACTUAL_PFY  idx 16)
    """

    # ── helpers ───────────────────────────────────────────────────────────────
>>>>>>> Stashed changes
    def safe_int(value):
        return int(value) if value is not None else 0

    def to_mins(seconds):
<<<<<<< Updated upstream
        mins = float(seconds) / 60.0 if seconds is not None else 0.0
        return int(mins) if mins % 1 == 0 else round(mins, 1)

    def to_hours(seconds):
        if seconds is None:
            return 0.0
        try:
            val = float(seconds)
            if val <= 24:
                return val
            hours = int(val // 3600)
            minutes = int((val % 3600) // 60)
            return round(hours + minutes / 100.0, 2)
        except (ValueError, TypeError):
            return 0.0

    def populate_row(ws, row_idx, query_row):
        if not query_row:
            return
        ws.cell(row=row_idx, column=5).value = safe_int(query_row[13])
        ws.cell(row=row_idx, column=6).value = safe_int(query_row[14])
        ws.cell(row=row_idx, column=7).value = safe_int(query_row[14]) - safe_int(query_row[13])
        
        ws.cell(row=row_idx, column=8).value = safe_int(query_row[4])
        ws.cell(row=row_idx, column=9).value = safe_int(query_row[6])
        ws.cell(row=row_idx, column=10).value = safe_int(query_row[8])
        ws.cell(row=row_idx, column=11).value = safe_int(query_row[10])
        ws.cell(row=row_idx, column=12).value = safe_int(query_row[12])
        
        ws.cell(row=row_idx, column=13).value = safe_int(query_row[1])
        ws.cell(row=row_idx, column=14).value = safe_int(query_row[2])
        ws.cell(row=row_idx, column=15).value = safe_int(query_row[2]) - safe_int(query_row[1])
        
        ws.cell(row=row_idx, column=16).value = safe_int(query_row[20])
        ws.cell(row=row_idx, column=18).value = safe_int(query_row[16])

        for col_idx in [5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 18]:
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.number_format = '@'

    try:
        # 1. Parse and validate dates
        try:
            report_date = datetime.datetime.strptime(payload.ReportDate, "%Y-%m-%d").date()
            start_date = datetime.datetime.strptime(payload.StartDate, "%Y-%m-%d").date()
            last_date = datetime.datetime.strptime(payload.LastDate, "%Y-%m-%d").date()
            # Define output path dynamically in the user's Downloads directory
            downloads_dir = Path.home() / "Downloads"
            if not downloads_dir.exists():
                downloads_dir = Path.home()
            now_str = datetime.datetime.now().strftime("%H-%M-%S")
            output_path = downloads_dir / f"MProductionReport_{report_date:%Y-%m-%d}_{now_str}.xlsx"
        except ValueError as val_err:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid date format. Expected YYYY-MM-DD. Error: {str(val_err)}"
            )


        # 2. Query Holidays
        cursor = conn.cursor()
      

        # 3. Run main SQL query for Micky
        cursor.execute(SaarthiMickyReportTCFBIW(payload.ReportDate, payload.StartDate, payload.LastDate, payload.Shift,"S_TCF"))
        rows = cursor.fetchall()
        
        columns = [col[0] for col in cursor.description]
        print("--- Daily Line Stop Row Details ---")
        for row in rows:
            for col_name, val in zip(columns, row):
                print(f"{col_name} : {val}")

        cursor.execute(SaarthiMickyReportTCFBIW(payload.ReportDate, payload.StartDate, payload.LastDate, payload.Shift,"M_TCF"))
        rows1 = cursor.fetchall()

        cursor.execute(SaarthiMickyReportTCFBIW(payload.ReportDate, payload.StartDate, payload.LastDate, payload.Shift,"S_BIW"))
        rows2 = cursor.fetchall()



        # 4. Run line stop queries (Daily & Monthly) and Production Loss query
        cursor.execute(LineStopRecordDaily(payload.ReportDate))
        daily_line_stops = cursor.fetchall()

       

        cursor.execute(LineStopRecordMonthly(payload.ReportDate))
        monthly_line_stops = cursor.fetchall()

  
  
        # 6. Load Excel template and populate sheets
        wb = openpyxl.load_workbook(TEMPLATE_PATH)

  

        # G. Update Date & Time and date formatting in summary sheets
        current_date = datetime.date.today()
        current_time = datetime.datetime.now().time()
        
        if "Prod Report" in wb.sheetnames:
            ws = wb["Prod Report"]
            
            # Update Date & Time
            ws.cell(row=2, column=2).value = current_date.strftime("%d-%m-%Y")
            ws.cell(row=3, column=2).value = current_time
            
            # Update Month under General Header (R2)
            ws.cell(row=2, column=18).value = report_date.strftime("%b %Y")
            
            # Available Working Days & Remaining Days
            days_in_month = calendar.monthrange(report_date.year, report_date.month)[1]
            remaining_days = days_in_month - 4  # 4 holidays subtracted
            ws.cell(row=3, column=18).value = remaining_days
            
            today_day = report_date.day
            remaining_days_left = max(0, days_in_month - today_day - 4)
            ws.cell(row=4, column=18).value = remaining_days_left
            
            # Update Month Label in sections (E8, E17, E25)
            month_label = report_date.strftime("%b %Y")
            for r in [8, 17, 25]:
                ws.cell(row=r, column=5).value = month_label
                ws.cell(row=r, column=5).number_format = '@'
                
            # Update Daily Date Label in sections (M8, M17, M25)
            daily_label = report_date.strftime("%d-%b-%Y")
            for r in [8, 17, 25]:
                ws.cell(row=r, column=13).value = daily_label
                ws.cell(row=r, column=13).number_format = '@'
                
            # Update YTD Year Header (P8, P17, P25)
            ytd_label = f"{(start_date.year) % 100:02d}-{(last_date.year) % 100:02d}"
            for r in [8, 17, 25]:
                ws.cell(row=r, column=16).value = ytd_label
                ws.cell(row=r, column=16).number_format = '@'
                
            # Update FY Year Header (R8, R17, R25)
            fy_label = f"{(start_date.year - 1) % 100:02d}-{(start_date.year) % 100:02d}"
            for r in [8, 17, 25]:
                ws.cell(row=r, column=18).value = fy_label
                ws.cell(row=r, column=18).number_format = '@'

            # Update Weekly Headers (W1, W2, W3, W4, W5 -> e.g. W23, W24, W25, W26, W27)
            first_of_month = datetime.date(report_date.year, report_date.month, 1)
            first_weekday = (first_of_month.weekday() + 1) % 7 + 1
            week_headers = {}
            for k in range(1, 6):
                d_start = max(1, 7 * (k - 1) - first_weekday + 2)
                d_end = min(days_in_month, 7 * k - first_weekday + 1)
                if d_start <= days_in_month:
                    mid_day = datetime.date(report_date.year, report_date.month, (d_start + d_end) // 2)
                    week_headers[k] = f"W{mid_day.isocalendar()[1]}"
                else:
                    week_headers[k] = ""

            for r in [9, 18, 26]:
                for k in range(1, 6):
                    col_idx = 7 + k  # columns 8, 9, 10, 11, 12
                    ws.cell(row=r, column=col_idx).value = week_headers[k]
                    ws.cell(row=r, column=col_idx).number_format = '@'

            # Populate Vehicle & BIW Rollout tables
            # Saarthi Vehicle: Row 10
            if rows:
                populate_row(ws, 10, rows[0])
            # Micky Vehicle: Row 11
            if rows1:
                populate_row(ws, 11, rows1[0])
            # Saarthi BIW: Row 19
            if rows2:
                populate_row(ws, 19, rows2[0])

            # A. Populate Daily Line Stops in 'Prod Report' sheet
            daily_row_mapping = {
                "Process Call": 35,
                "Material Call": 36,
                "Quality Call": 37,
                "Maintenance Call": 38,
                "Other": 39
            }
            
            # Initialize table cells to 0/empty to ensure old/cached data is cleared
            for r in daily_row_mapping.values():
                for c in [5, 6, 7, 8, 9, 10, 11, 13]: # Chassis, Trim, Saarthi Main, Saarthi Sub, I-PUMA Main, I-PUMA Sub, Cargo Main, Cargo Sub
                    ws.cell(row=r, column=c).value = 0
                ws.cell(row=r, column=15).value = "" # Chassis Line Loss Reason
                ws.cell(row=r, column=19).value = "" # Remark
            
            # Date at E32
            ws.cell(row=32, column=5).value = report_date.strftime("%d-%m-%Y")
            
            for row in daily_line_stops:
                type_of_call = row[0]
                r = daily_row_mapping.get(type_of_call)
                if r:
                    val_chassis = to_mins(row[1])
                    val_trim = to_mins(row[2])
                    val_saarthi_main = to_mins(row[3])
                    val_saarthi_sub = to_mins(row[4])
                    val_ipuma_main = to_mins(row[5])
                    val_ipuma_sub = to_mins(row[6])
                    val_cargo_main = to_mins(row[7])
                    val_cargo_sub = to_mins(row[8])
                    
                    ws.cell(row=r, column=5).value = val_chassis
                    ws.cell(row=r, column=6).value = val_trim
                    ws.cell(row=r, column=7).value = val_saarthi_main
                    ws.cell(row=r, column=8).value = val_saarthi_sub
                    ws.cell(row=r, column=9).value = val_ipuma_main
                    ws.cell(row=r, column=10).value = val_ipuma_sub
                    ws.cell(row=r, column=11).value = val_cargo_main
                    ws.cell(row=r, column=13).value = val_cargo_sub
                    
                    ws.cell(row=r, column=15).value = row[9]
                    ws.cell(row=r, column=19).value = row[10]

        # Populate Daily sheet
=======
        if seconds is None:
            return 0
        mins = float(seconds) / 60.0
        return int(mins) if mins % 1 == 0 else round(mins, 1)

    try:
        # ── 1. Parse dates ────────────────────────────────────────────────────
        try:
            report_date = datetime.datetime.strptime(payload.ReportDate, "%Y-%m-%d").date()
            start_date  = datetime.datetime.strptime(payload.StartDate,  "%Y-%m-%d").date()
            last_date   = datetime.datetime.strptime(payload.LastDate,   "%Y-%m-%d").date()

            downloads_dir = Path.home() / "Downloads"
            if not downloads_dir.exists():
                downloads_dir = Path.home()
            now_str     = datetime.datetime.now().strftime("%H-%M-%S")
            output_path = downloads_dir / f"ProductionReport_R3_{report_date:%Y-%m-%d}_{now_str}.xlsx"
        except ValueError as e:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid date format. Expected YYYY-MM-DD. Error: {e}",
            )

        # ── 2. Compute ISO weeks for the report month ─────────────────────────
        month_iso_weeks = get_month_iso_weeks(report_date)
        # month_iso_weeks e.g. [23, 24, 25, 26, 27] for June 2026
        # Cap at 5 columns (H-L); if a month has 6 ISO weeks, the 6th overflows
        display_weeks = month_iso_weeks[:5]
        print(f"[ProdReportType2] Report month ISO weeks: {month_iso_weeks} (displaying: {display_weeks})")

        # ── 3. Run SQL queries ────────────────────────────────────────────────
        cursor = conn.cursor()

        def run_aggregate(db_name: str):
            """Run SaarthiMickyReportTCFBIW1 (monthly / daily / YTD / FY)."""
            cursor.execute(SaarthiMickyReportTCFBIW1(
                payload.ReportDate, payload.StartDate, payload.LastDate,
                payload.Shift, db_name,
            ))
            rows = cursor.fetchall()
            print(f"  [aggregate] {db_name}: {len(rows)} row(s)")
            return rows

        def run_weekwise(db_name: str):
            """Run SaarthiMickyWeekwiseMonthly (ISO week breakdown)."""
            cursor.execute(SaarthiMickyWeekwiseMonthly(
                payload.ReportDate, payload.Shift, db_name,
            ))
            rows = cursor.fetchall()
            wkdict = build_weekwise_dict(rows)
            print(f"  [weekwise]  {db_name}: {wkdict}")
            return wkdict

        # Vehicle TCF
        rows_stcf    = run_aggregate("S_TCF")
        wkdict_stcf  = run_weekwise("S_TCF")

        rows_mtcf    = run_aggregate("M_TCF")
        wkdict_mtcf  = run_weekwise("M_TCF")

        # BIW
        rows_sbiw    = run_aggregate("S_BIW")
        wkdict_sbiw  = run_weekwise("S_BIW")

        # Station Call
        cursor.execute(LineStopRecordDaily(payload.ReportDate))
        daily_line_stops = cursor.fetchall()

        cursor.close()

        # ── 4. Load Excel template ────────────────────────────────────────────
        wb = openpyxl.load_workbook(TEMPLATE_PATH)

        if "Prod Report" not in wb.sheetnames:
            raise HTTPException(
                status_code=500,
                detail="Template sheet 'Prod Report' not found in ProductionReport_R3.xlsx",
            )
        ws = wb["Prod Report"]

        # ── 5. Report header (rows 2-4) ───────────────────────────────────────
        current_time = datetime.datetime.now().time()
        ws.cell(row=2, column=2).value = report_date.strftime("%d-%m-%Y")  # B2 Date
        ws.cell(row=3, column=2).value = current_time                       # B3 Time

        month_label = report_date.strftime("%b %Y")          # "Jun 2026"
        date_label  = report_date.strftime("%d-%b-%Y")       # "09-Jun-2026"
        fy_label    = (f"{start_date.year % 100:02d}-"
                       f"{last_date.year % 100:02d}")        # "26-27"

        # Working days from DAYS_IN_MONTH
        if rows_stcf:
            total_days     = safe_int(rows_stcf[0][IDX_DAYS_IN_MONTH])
            remaining_days = max(0, total_days - report_date.day - 4)
        else:
            total_days = remaining_days = 0

        ws.cell(row=2, column=18).number_format = "@"
        ws.cell(row=2, column=18).value = month_label    # R2
        ws.cell(row=3, column=18).value = total_days     # R3
        ws.cell(row=4, column=18).value = remaining_days # R4

        # ── 6. Section date/label headers & dynamic week labels ───────────────
        # header_row -> (label_row_for_sub_headers)
        SECTION_ROWS = {
            8:  9,    # Vehicle  -- section header row 8, sub-header (week labels) row 9
            17: 18,   # BIW      -- section header row 17, sub-header row 18
            25: 26,   # Paint    -- section header row 25, sub-header row 26
        }

        for hdr_row, sub_row in SECTION_ROWS.items():
            # Section-level labels (month, daily date, YTD FY)
            ws.cell(row=hdr_row, column=5).number_format  = "@"
            ws.cell(row=hdr_row, column=5).value  = month_label   # E: Monthly label
            ws.cell(row=hdr_row, column=13).number_format = "@"
            ws.cell(row=hdr_row, column=13).value = date_label    # M: Daily date
            ws.cell(row=hdr_row, column=16).number_format = "@"
            ws.cell(row=hdr_row, column=16).value = fy_label      # P: YTD FY label

            # Dynamic ISO week sub-headers (e.g. W23, W24, W25, W26, W27)
            for slot_idx, col in enumerate(WEEK_EXCEL_COLS):
                if slot_idx < len(display_weeks):
                    ws.cell(row=sub_row, column=col).value = f"W{display_weeks[slot_idx]}"
                else:
                    ws.cell(row=sub_row, column=col).value = ""   # clear unused slot

        # ── 7. Helper: write one vehicle / BIW row ────────────────────────────
        def write_vehicle_row(
            row_num: int,
            agg_rows,          # from SaarthiMickyReportTCFBIW1
            wkdict: dict,      # from SaarthiMickyWeekwiseMonthly
        ):
            """
            Fill E-R for one production row using aggregate + weekwise data.
            wkdict: {iso_week: (wk_target, wk_actual)}
            """
            if agg_rows:
                r = agg_rows[0]
                monthly_tgt = safe_int(r[IDX_MONTHLY_TARGET])
                monthly_act = safe_int(r[IDX_MONTHLY_ACTUAL])
                daily_tgt   = safe_int(r[IDX_DAILY_TARGET])
                daily_act   = safe_int(r[IDX_DAILY_ACTUAL])
                ytd_act     = safe_int(r[IDX_YTD_ACTUAL])
                pfy_act     = safe_int(r[IDX_PFY_ACTUAL])
            else:
                monthly_tgt = monthly_act = daily_tgt = daily_act = 0
                ytd_act = pfy_act = 0

            # Monthly (E, F, G)
            ws.cell(row=row_num, column=5).value  = monthly_tgt
            ws.cell(row=row_num, column=6).value  = monthly_act
            ws.cell(row=row_num, column=7).value  = monthly_act - monthly_tgt

            # Dynamic ISO-week actuals (H-L) — up to 5 weeks
            for slot_idx, col in enumerate(WEEK_EXCEL_COLS):
                if slot_idx < len(display_weeks):
                    iso_wk = display_weeks[slot_idx]
                    _, wk_act = wkdict.get(iso_wk, (0, 0))
                    ws.cell(row=row_num, column=col).value = wk_act
                else:
                    ws.cell(row=row_num, column=col).value = 0  # clear unused slot

            # Daily (M, N, O)
            ws.cell(row=row_num, column=13).value = daily_tgt
            ws.cell(row=row_num, column=14).value = daily_act
            ws.cell(row=row_num, column=15).value = daily_act - daily_tgt

            # YTD (P) and Previous FY (R)
            ws.cell(row=row_num, column=16).value = ytd_act
            ws.cell(row=row_num, column=18).value = pfy_act

        # ── 8. Vehicle Production Details (rows 10-13) ───────────────────────
        write_vehicle_row(10, rows_stcf, wkdict_stcf)   # Saarthi <- S_TCF
        write_vehicle_row(11, rows_mtcf, wkdict_mtcf)   # Micky   <- M_TCF
        write_vehicle_row(12, None,      {})             # Cargo   -> 0
        write_vehicle_row(13, None,      {})             # I-Puma  -> 0

        # ── 9. BIW Production Details (rows 19-21) ───────────────────────────
        write_vehicle_row(19, rows_sbiw, wkdict_sbiw)   # Saarthi <- S_BIW
        write_vehicle_row(20, None,      {})             # Cargo   -> 0
        write_vehicle_row(21, None,      {})             # I-Puma  -> 0

        # ── 10. Paint Shop (rows 27-30) -- all 0, Paint MTD Rollout empty ────
        for paint_row in [27, 28, 29, 30]:
            write_vehicle_row(paint_row, None, {})

        # ── 11. Station Call section ──────────────────────────────────────────
        ws.cell(row=32, column=5).value = datetime.datetime.combine(
            report_date, datetime.time()
        )

        CALL_TYPE_ROWS = {
            "Process Call":     35,
            "Material Call":    36,
            "Quality Call":     37,
            "Maintenance Call": 38,
            "Other":            39,
        }

        line_stop_map: dict = {}
        for ls_row in daily_line_stops:
            ct = ls_row[0]
            if ct in CALL_TYPE_ROWS:
                line_stop_map[ct] = ls_row

        for call_text, excel_row in CALL_TYPE_ROWS.items():
            ls = line_stop_map.get(call_text)
            if ls:
                chassis      = to_mins(ls[1])
                trim         = to_mins(ls[2])
                saarthi_main = to_mins(ls[3])
                saarthi_sub  = to_mins(ls[4])
                ipuma_main   = to_mins(ls[5])
                ipuma_sub    = to_mins(ls[6])
                cargo_main   = to_mins(ls[7])
                cargo_sub    = to_mins(ls[8])
                loss_reason  = ls[9]  if ls[9]  is not None else ""
                remark       = ls[10] if ls[10] is not None else ""
            else:
                chassis = trim = saarthi_main = saarthi_sub = 0
                ipuma_main = ipuma_sub = cargo_main = cargo_sub = 0
                loss_reason = remark = ""

            ws.cell(row=excel_row, column=5).value  = chassis
            ws.cell(row=excel_row, column=6).value  = trim
            ws.cell(row=excel_row, column=7).value  = saarthi_main
            ws.cell(row=excel_row, column=8).value  = saarthi_sub
            ws.cell(row=excel_row, column=9).value  = ipuma_main
            ws.cell(row=excel_row, column=10).value = ipuma_sub
            ws.cell(row=excel_row, column=11).value = cargo_main
            ws.cell(row=excel_row, column=13).value = cargo_sub
            ws.cell(row=excel_row, column=15).value = loss_reason
            ws.cell(row=excel_row, column=19).value = remark

        # ── 12. Daily sheet (if present) ─────────────────────────────────────
>>>>>>> Stashed changes
        if "Daily" in wb.sheetnames:
            ws_daily = wb["Daily"]

            def write_daily_block(data_row: int, agg_rows):
                if not agg_rows:
<<<<<<< Updated upstream
                    # Clear placeholders to 0/empty to prevent raw templates from being left behind
                    ws_daily.cell(row=data_row, column=3).value  = report_date.strftime("%d-%m-%Y")
                    ws_daily.cell(row=data_row, column=4).value  = 0
                    ws_daily.cell(row=data_row, column=5).value  = 0
                    ws_daily.cell(row=data_row, column=6).value  = 0
                    ws_daily.cell(row=data_row, column=7).value  = 0
                    ws_daily.cell(row=data_row, column=8).value  = 0
                    ws_daily.cell(row=data_row, column=9).value  = 0
                    ws_daily.cell(row=data_row, column=10).value = 0
                    ws_daily.cell(row=data_row, column=11).value = 0
                    ws_daily.cell(row=data_row, column=12).value = remaining_days
                    ws_daily.cell(row=data_row, column=13).value = remaining_days_left
                    return
                
                r = agg_rows[0]
                daily_tgt   = safe_int(r[1])
                daily_act   = safe_int(r[2])
                monthly_tgt = safe_int(r[13])
                monthly_act = safe_int(r[14])
=======
                    return
                r = agg_rows[0]
                daily_tgt   = safe_int(r[IDX_DAILY_TARGET])
                daily_act   = safe_int(r[IDX_DAILY_ACTUAL])
                monthly_tgt = safe_int(r[IDX_MONTHLY_TARGET])
                monthly_act = safe_int(r[IDX_MONTHLY_ACTUAL])
>>>>>>> Stashed changes

                ws_daily.cell(row=data_row, column=3).value  = report_date.strftime("%d-%m-%Y")
                ws_daily.cell(row=data_row, column=4).value  = daily_tgt
                ws_daily.cell(row=data_row, column=5).value  = daily_act
                ws_daily.cell(row=data_row, column=6).value  = daily_act - daily_tgt
                ws_daily.cell(row=data_row, column=7).value  = monthly_tgt
                ws_daily.cell(row=data_row, column=8).value  = monthly_act
                ws_daily.cell(row=data_row, column=9).value  = monthly_act - monthly_tgt
<<<<<<< Updated upstream
                ws_daily.cell(row=data_row, column=10).value = safe_int(r[20])
                ws_daily.cell(row=data_row, column=11).value = safe_int(r[16])
                ws_daily.cell(row=data_row, column=12).value = remaining_days
                ws_daily.cell(row=data_row, column=13).value = remaining_days_left

            write_daily_block(data_row=2,  agg_rows=rows1)  # M_TCF
            write_daily_block(data_row=12, agg_rows=rows)   # S_TCF
            write_daily_block(data_row=21, agg_rows=None)   # M_BIW
            write_daily_block(data_row=31, agg_rows=rows2)  # S_BIW

         
        # 7. Save and return workbook
        wb.save(output_path)

        return {
            "status": "success",
            "filepath": str(output_path),
            "filename": output_path.name
=======
                ws_daily.cell(row=data_row, column=10).value = safe_int(r[IDX_YTD_ACTUAL])
                ws_daily.cell(row=data_row, column=11).value = safe_int(r[IDX_PFY_ACTUAL])
                ws_daily.cell(row=data_row, column=12).value = total_days
                ws_daily.cell(row=data_row, column=13).value = remaining_days

            write_daily_block(data_row=2,  agg_rows=rows_mtcf)
            write_daily_block(data_row=12, agg_rows=rows_stcf)
            write_daily_block(data_row=31, agg_rows=rows_sbiw)

        # ── 13. Save and return ───────────────────────────────────────────────
        wb.save(output_path)
        print(f"[ProdReportType2] Saved -> {output_path}")

        return {
            "status":      "success",
            "filepath":    str(output_path),
            "filename":    output_path.name,
            "weekLabels":  [f"W{w}" for w in display_weeks],  # e.g. ["W23","W24","W25","W26","W27"]
>>>>>>> Stashed changes
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


<<<<<<< Updated upstream
@router.post("/prod-report-type2/open-file")
def open_file_endpoint(payload: OpenFileRequest):
=======
# ── Utility endpoints ─────────────────────────────────────────────────────────

@router.post("/prod-report-type2/open-file")
def open_report_type2_file(payload: OpenFileRequest):
    """Open a generated ProductionReport_R3 file on the host OS."""
>>>>>>> Stashed changes
    try:
        path = os.path.abspath(payload.filepath)
        if not os.path.exists(path):
            raise HTTPException(status_code=404, detail="File not found")
<<<<<<< Updated upstream
        
=======
>>>>>>> Stashed changes
        if os.name == "nt":
            os.startfile(path)
        else:
            opener = "open" if sys.platform == "darwin" else "xdg-open"
            subprocess.call([opener, path])
<<<<<<< Updated upstream
            
=======
>>>>>>> Stashed changes
        return {"status": "success", "message": "File opened successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/prod-report-type2/open-folder")
<<<<<<< Updated upstream
def open_folder_endpoint(payload: OpenFileRequest):
=======
def open_report_type2_folder(payload: OpenFileRequest):
    """Highlight a generated ProductionReport_R3 file in Windows Explorer."""
>>>>>>> Stashed changes
    try:
        path = os.path.abspath(payload.filepath)
        if not os.path.exists(path):
            raise HTTPException(status_code=404, detail="File not found")
<<<<<<< Updated upstream
        
        if os.name == "nt":
            # Highlight the file in Explorer
            subprocess.Popen(f'explorer /select,"{path}"')

        return {"status": "success", "message": "Folder opened successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
=======
        if os.name == "nt":
            subprocess.Popen(f'explorer /select,"{path}"')
        return {"status": "success", "message": "Folder opened successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
>>>>>>> Stashed changes
