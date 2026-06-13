"""
Chassis Loss Report — FastAPI Router
=====================================
Generates an Excel report (based on Chasis_Loss_Report.xlsx template) populated with:

  Summary block (A7:M12)
  ├── One row per chassis station (CH-10 … CH-60, stations 10-60 mapped to CH_S1-CH_S6)
  └── Columns: Station No, Total Delay (Min/Freq), Process (Min/Freq),
               Material (Min/Freq), Quality (Min/Freq), Maintenance (Min/Freq), Other (Min/Freq)
  Note: All time values are converted from SECONDS to MINUTES.

  Station Call History (row 22 onwards)
  └── All raw records from CH_S1–CH_S6 for the requested time window, ordered by DT ASC.

Request params: StartTime, EndTime  (format: YYYY-MM-DD HH:MM:SS)
"""
import os
import sys
import subprocess
import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Border, Side
import pyodbc
from fastapi import APIRouter, Depends, HTTPException

from app.core.database import get_db_connection
from app.schemas.chassis_loss_report import ChassisLossReportRequest, OpenChassisFileRequest
from app.constant.chassis_loss_queries import ChassisLossReportSummary, ChassisLossReportHistory

router = APIRouter()

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parents[3]
TEMPLATE_PATH = BASE_DIR / "Chasis_Loss_Report.xlsx"

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _to_mins(seconds) -> float:
    """Convert seconds to minutes, rounded to 1 decimal place.
    Returns 0.0 for None / falsy values.
    """
    if seconds is None:
        return 0.0
    try:
        val = float(seconds)
        if val == 0.0:
            return 0.0
        mins = val / 60.0
        # Return as int if no decimal part, else 1 dp
        return int(mins) if mins % 1 == 0 else round(mins, 1)
    except (ValueError, TypeError):
        return 0.0


def _fmt_dt(dt_val) -> str:
    """Format a datetime-like value to a readable string, or return empty string."""
    if dt_val is None:
        return ""
    try:
        if hasattr(dt_val, "strftime"):
            return dt_val.strftime("%Y-%m-%d %H:%M:%S")
        return str(dt_val)
    except Exception:
        return str(dt_val)


def _fmt_time(dt_val) -> str:
    """Extract time portion HH:MM:SS from a datetime."""
    if dt_val is None:
        return ""
    try:
        if hasattr(dt_val, "strftime"):
            return dt_val.strftime("%H:%M:%S")
        return str(dt_val)
    except Exception:
        return str(dt_val)


def _get_sheet(workbook, name):
    """Return a sheet with a patched .cell() that skips MergedCells transparently."""
    if name not in workbook.sheetnames:
        return None
    sheet = workbook[name]
    orig_cell = sheet.cell

    def safe_cell(*args, **kwargs):
        cell = orig_cell(*args, **kwargs)
        if type(cell).__name__ == "MergedCell":
            for r in sheet.merged_cells.ranges:
                if cell.coordinate in r:
                    return orig_cell(row=r.min_row, column=r.min_col)
        return cell

    sheet.cell = safe_cell
    return sheet


# ---------------------------------------------------------------------------
# Station summary — column layout in the Excel summary block
# ---------------------------------------------------------------------------
#
# Row 7  -> Station CH-10 (CH_S1)
# Row 8  -> Station CH-20 (CH_S2)
# Row 9  -> Station CH-30 (CH_S3)
# Row 10 -> Station CH-40 (CH_S4)
# Row 11 -> Station CH-50 (CH_S5)
# Row 12 -> Station CH-60 (CH_S6)
#
# Col 1  (A) = StationLabel text (e.g. "10")
# Col 2  (B) = Total Delay   Min  (TotalMin  → seconds → minutes)
# Col 3  (C) = Total Delay   Freq (TotalFreq)
# Col 4  (D) = Process Call  Min  (ProcessMin → minutes)
# Col 5  (E) = Process Call  Freq (ProcessFreq)
# Col 6  (F) = Material Call Min  (MaterialMin → minutes)
# Col 7  (G) = Material Call Freq (MaterialFreq)
# Col 8  (H) = Quality Call  Min  (QualityMin → minutes)
# Col 9  (I) = Quality Call  Freq (QualityFreq)
# Col 10 (J) = Maint Call    Min  (MaintMin → minutes)
# Col 11 (K) = Maint Call    Freq (MaintFreq)
# Col 12 (L) = Other Call    Min  (OtherMin → minutes)
# Col 13 (M) = Other Call    Freq (OtherFreq)
#
STATION_ROW_MAP = {
    "CH-10": 7,
    "CH-20": 8,
    "CH-30": 9,
    "CH-40": 10,
    "CH-50": 11,
    "CH-60": 12,
}

# Station number display labels (just the numeric part shown in template)
STATION_DISPLAY = {
    "CH-10": "10",
    "CH-20": "20",
    "CH-30": "30",
    "CH-40": "40",
    "CH-50": "50",
    "CH-60": "60",
}


# ---------------------------------------------------------------------------
# Endpoint: Generate Report
# ---------------------------------------------------------------------------
@router.post("/chassis-loss-report")
def generate_chassis_loss_report(
    payload: ChassisLossReportRequest,
    conn: pyodbc.Connection = Depends(get_db_connection),
):
    """
    Generates the Chassis Loss Report Excel file from the Chasis_Loss_Report.xlsx template.

    Populates:
    - Summary block (rows 7-12, cols A-M): per-station aggregated loss in minutes + frequency
    - Metadata: start/end time, generated date/time
    - Station Call History (rows 22+): all raw CH_Sx records in the time window

    Parameters
    ----------
    StartTime : str
        Start of the reporting window, e.g. "2026-06-13 06:00:00"
    EndTime : str
        End of the reporting window, e.g. "2026-06-13 18:00:00"
    """
    try:
        # ------------------------------------------------------------------
        # 1. Validate the template file exists
        # ------------------------------------------------------------------
        if not TEMPLATE_PATH.exists():
            raise HTTPException(
                status_code=500,
                detail=f"Template not found: {TEMPLATE_PATH}",
            )

        # ------------------------------------------------------------------
        # 2. Parse & validate datetime params
        # ------------------------------------------------------------------
        fmt = "%Y-%m-%d %H:%M:%S"
        try:
            start_dt = datetime.datetime.strptime(payload.StartTime, fmt)
            end_dt = datetime.datetime.strptime(payload.EndTime, fmt)
        except ValueError as e:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid datetime format. Expected 'YYYY-MM-DD HH:MM:SS'. Error: {e}",
            )

        if end_dt <= start_dt:
            raise HTTPException(
                status_code=400,
                detail="EndTime must be after StartTime.",
            )

        # ------------------------------------------------------------------
        # 3. Build output path
        # ------------------------------------------------------------------
        downloads_dir = Path.home() / "Downloads"
        if not downloads_dir.exists():
            downloads_dir = Path.home()
        now_str = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        output_path = downloads_dir / f"ChassisLossReport_{now_str}.xlsx"

        # ------------------------------------------------------------------
        # 4. Run SQL queries
        # ------------------------------------------------------------------
        cursor = conn.cursor()

        # 4a. Summary query — per-station aggregated data
        cursor.execute(ChassisLossReportSummary(payload.StartTime, payload.EndTime))
        summary_rows = cursor.fetchall()

        # Build a dict keyed by StationLabel for easy lookup
        # Columns from query:
        #   0:StationLabel, 1:TotalMin, 2:TotalFreq,
        #   3:ProcessMin, 4:ProcessFreq, 5:MaterialMin, 6:MaterialFreq,
        #   7:QualityMin, 8:QualityFreq, 9:MaintMin, 10:MaintFreq,
        #   11:OtherMin, 12:OtherFreq
        summary_data = {}
        for row in summary_rows:
            summary_data[row[0]] = row

        # 4b. History query — all raw records
        cursor.execute(ChassisLossReportHistory(payload.StartTime, payload.EndTime))
        history_rows = cursor.fetchall()
        # Columns:
        #   0:DT, 1:StationLabel, 2:StationNo, 3:Typeofcall, 4:CallTypeText,
        #   5:LossTime(sec), 6:CallEndTime, 7:Reason, 8:Remark

        # ------------------------------------------------------------------
        # 5. Load Excel template
        # ------------------------------------------------------------------
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        ws = _get_sheet(wb, "Sheet1")

        if ws is None:
            raise HTTPException(
                status_code=500,
                detail="Sheet 'Sheet1' not found in the template.",
            )

        # ------------------------------------------------------------------
        # 6. Populate metadata (rows 2-3)
        # ------------------------------------------------------------------
        current_dt = datetime.datetime.now()

        # Row 2: Month in G2 area — use start_dt month
        ws.cell(row=2, column=7).value = start_dt.strftime("%Y-%m-01")

        # Row 2: DATE in J2 area — current date
        ws.cell(row=2, column=9).value = "DATE"
        ws.cell(row=2, column=10).value = current_dt.strftime("%Y-%m-%d")

        # Row 3: Report Time range
        ws.cell(row=3, column=1).value = (
            f"Report Time: {start_dt.strftime('%Y-%m-%d %H:%M:%S')} "
            f"to {end_dt.strftime('%Y-%m-%d %H:%M:%S')}"
        )
        # TIME in J3
        ws.cell(row=3, column=10).value = current_dt.strftime("%I:%M%p")

        # ------------------------------------------------------------------
        # 7. Populate summary block (rows 7-12, cols A-M)
        # ------------------------------------------------------------------
        for station_label, excel_row in STATION_ROW_MAP.items():
            # Col A (1): Station display number
            ws.cell(row=excel_row, column=1).value = STATION_DISPLAY[station_label]

            if station_label in summary_data:
                row = summary_data[station_label]

                # Col B (2): Total Delay Min (seconds → minutes)
                ws.cell(row=excel_row, column=2).value = _to_mins(row[1])
                # Col C (3): Total Delay Freq
                ws.cell(row=excel_row, column=3).value = int(row[2]) if row[2] else 0

                # Col D (4): Process Call Min
                ws.cell(row=excel_row, column=4).value = _to_mins(row[3])
                # Col E (5): Process Call Freq
                ws.cell(row=excel_row, column=5).value = int(row[4]) if row[4] else 0

                # Col F (6): Material Call Min
                ws.cell(row=excel_row, column=6).value = _to_mins(row[5])
                # Col G (7): Material Call Freq
                ws.cell(row=excel_row, column=7).value = int(row[6]) if row[6] else 0

                # Col H (8): Quality Call Min
                ws.cell(row=excel_row, column=8).value = _to_mins(row[7])
                # Col I (9): Quality Call Freq
                ws.cell(row=excel_row, column=9).value = int(row[8]) if row[8] else 0

                # Col J (10): Maintenance Call Min
                ws.cell(row=excel_row, column=10).value = _to_mins(row[9])
                # Col K (11): Maintenance Call Freq
                ws.cell(row=excel_row, column=11).value = int(row[10]) if row[10] else 0

                # Col L (12): Other Call Min
                ws.cell(row=excel_row, column=12).value = _to_mins(row[11])
                # Col M (13): Other Call Freq
                ws.cell(row=excel_row, column=13).value = int(row[12]) if row[12] else 0

            else:
                # Station has no data in the window — fill zeros
                for col in range(2, 14):
                    ws.cell(row=excel_row, column=col).value = 0

        # ------------------------------------------------------------------
        # 8. Clear existing history rows (row 15 onwards) then repopulate
        # ------------------------------------------------------------------
        # Clear any pre-existing data starting from row 15
        max_row = ws.max_row
        empty_border = Border()
        for r in range(15, max_row + 1):
            for c in range(1, 7):
                try:
                    cell = ws.cell(row=r, column=c)
                    cell.value = None
                    cell.border = empty_border
                except Exception:
                    pass

        # ------------------------------------------------------------------
        # 9. Populate Station Call History (row 15 onwards)
        #    Header already present in template.
        # ------------------------------------------------------------------
        history_start_row = 15
        thin_side = Side(style="thin", color="000000")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        for idx, h_row in enumerate(history_rows):
            excel_r = history_start_row + idx

            # Col A (1): DT - call start datetime
            ws.cell(row=excel_r, column=1).value = _fmt_dt(h_row[0])

            # Col B (2): Station No (numeric, e.g. 10, 20...)
            ws.cell(row=excel_r, column=2).value = h_row[2]  # StationNo

            # Col C (3): Call Type text
            ws.cell(row=excel_r, column=3).value = h_row[4]  # CallTypeText

            # Col E (5): Call Time in minutes (LossTime seconds → minutes)
            ws.cell(row=excel_r, column=4).value = _to_mins(h_row[5])  # LossTime

            # Col G (7): Call End Time
            ws.cell(row=excel_r, column=5).value = _fmt_time(h_row[6])  # CallEndTime

            # Col I (9): Station Call Reason (Reason field)
            ws.cell(row=excel_r, column=6).value = str(h_row[7]) if h_row[7] is not None else ""

            # Col L (12): Remark
            ws.cell(row=excel_r, column=7).value = str(h_row[8]) if h_row[8] is not None else ""

            # Set thin borders for columns 1 to 13 (A to M) to enclose the data row
            for c in range(1, 7):
                ws.cell(row=excel_r, column=c).border = thin_border

        # ------------------------------------------------------------------
        # 10. Save and return
        # ------------------------------------------------------------------
        wb.save(output_path)

        print(
            f"[ChassisLossReport] Generated: {output_path} | "
            f"Summary rows: {len(summary_rows)} | History rows: {len(history_rows)}"
        )

        return {
            "status": "success",
            "filepath": str(output_path),
            "filename": output_path.name,
            "summary_stations": len(summary_rows),
            "history_records": len(history_rows),
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# Endpoint: Open file in Explorer/Finder
# ---------------------------------------------------------------------------
@router.post("/chassis-loss-report/open-file")
def open_chassis_file(payload: OpenChassisFileRequest):
    """Open the generated report file using the OS default application."""
    try:
        path = os.path.abspath(payload.filepath)
        if not os.path.exists(path):
            raise HTTPException(status_code=404, detail="File not found")

        if os.name == "nt":
            os.startfile(path)
        else:
            opener = "open" if sys.platform == "darwin" else "xdg-open"
            subprocess.call([opener, path])

        return {"status": "success", "message": "File opened successfully"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# Endpoint: Open containing folder
# ---------------------------------------------------------------------------
@router.post("/chassis-loss-report/open-folder")
def open_chassis_folder(payload: OpenChassisFileRequest):
    """Highlight the generated report file in Windows Explorer."""
    try:
        path = os.path.abspath(payload.filepath)
        if not os.path.exists(path):
            raise HTTPException(status_code=404, detail="File not found")

        if os.name == "nt":
            subprocess.Popen(f'explorer /select,"{path}"')

        return {"status": "success", "message": "Folder opened successfully"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
