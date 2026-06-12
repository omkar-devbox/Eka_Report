import os
import sys
import subprocess
import calendar
import datetime
from datetime import date
import openpyxl
import pyodbc
from pathlib import Path
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse

from app.core.database import get_db_connection
from app.schemas.mgmt_prod_report import MgmtProdReportRequest, OpenFileRequest
from app.constant.queris import (
    SaarthiMickyReportTCFBIW,
    LineStopRecordDaily,
    LineStopRecordMonthly,
    ProductionLossDaily,
    ChassisLineStatus
)

router = APIRouter()

BASE_DIR = Path(__file__).resolve().parents[3]
TEMPLATE_PATH = BASE_DIR / "MProductionReport.xlsx"
# OUTPUT_PATH will be defined dynamically after parsing dates


@router.post("/mgmt-production-report")
def generate_mgmt_production_report(
    payload: MgmtProdReportRequest,
    conn: pyodbc.Connection = Depends(get_db_connection)
):
    """
    Generates and populates the Management Production Report for Micky (M_TCF).
    Reads date parameters dynamically from request body, queries raw daily log data,
    performs Daily, Weekly, Monthly, MTD, YTD, and FY calculations, updates the Excel spreadsheet,
    and returns the updated report as a downloadable file.
    """
    # Helper to safely convert None to 0
    def safe_int(value):
        return int(value) if value is not None else 0

    def to_mins(seconds):
        mins = float(seconds) / 60.0 if seconds is not None else 0.0
        return int(mins) if mins % 1 == 0 else round(mins, 1)

    def to_hours(seconds):
        if seconds is None:
            return 0.0
        try:
            val = float(seconds)
            if val <= 24:
                return val
            hours = int(val // 3600)
            minutes = int((val % 3600) // 60)
            return round(hours + minutes / 100.0, 2)
        except (ValueError, TypeError):
            return 0.0

    try:
        # 1. Parse and validate dates
        try:
            report_date = datetime.datetime.strptime(payload.ReportDate, "%Y-%m-%d").date()
            start_date = datetime.datetime.strptime(payload.StartDate, "%Y-%m-%d").date()
            last_date = datetime.datetime.strptime(payload.LastDate, "%Y-%m-%d").date()
            # Define output path dynamically in the user's Downloads directory
            downloads_dir = Path.home() / "Downloads"
            if not downloads_dir.exists():
                downloads_dir = Path.home()
            now_str = datetime.datetime.now().strftime("%H-%M-%S")
            output_path = downloads_dir / f"MProductionReport_{report_date:%Y-%m-%d}_{now_str}.xlsx"
        except ValueError as val_err:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid date format. Expected YYYY-MM-DD. Error: {str(val_err)}"
            )

        # Calculate 6 consecutive Yearly Weeks for the month starting from the first week (G9 to L9, columns 7 to 12)
        first_day = report_date.replace(day=1)
        first_monday = first_day - datetime.timedelta(days=first_day.weekday())
        weekly_headers = []
        for i in range(6):
            week_date = first_monday + datetime.timedelta(days=7 * i)
            iso_year, iso_week, iso_weekday = week_date.isocalendar()
            weekly_headers.append(f"W{iso_week}")


        # 2. Query Holidays
        cursor = conn.cursor()
      

        # 3. Run main SQL query for Micky
        cursor.execute(SaarthiMickyReportTCFBIW(payload.ReportDate, payload.StartDate, payload.LastDate, payload.Shift,"S_TCF"))
        rows = cursor.fetchall()
        
        columns = [col[0] for col in cursor.description]
        print("--- Daily Line Stop Row Details ---")
        for row in rows:
            for col_name, val in zip(columns, row):
                print(f"{col_name} : {val}")

        cursor.execute(SaarthiMickyReportTCFBIW(payload.ReportDate, payload.StartDate, payload.LastDate, payload.Shift,"M_TCF"))
        rows1 = cursor.fetchall()

        cursor.execute(SaarthiMickyReportTCFBIW(payload.ReportDate, payload.StartDate, payload.LastDate, payload.Shift,"S_BIW"))
        rows2 = cursor.fetchall()

        cursor.execute(SaarthiMickyReportTCFBIW(payload.ReportDate, payload.StartDate, payload.LastDate, payload.Shift,"M_BIW"))
        rows3 = cursor.fetchall()

        # 4. Run line stop queries (Daily & Monthly) and Production Loss query
        cursor.execute(LineStopRecordDaily(payload.ReportDate))
        daily_line_stops = cursor.fetchall()

       

        cursor.execute(LineStopRecordMonthly(payload.ReportDate))
        monthly_line_stops = cursor.fetchall()

        cursor.execute(ProductionLossDaily(payload.ReportDate, payload.Shift))
        daily_prod_loss = cursor.fetchall()

        cursor.execute(ChassisLineStatus(payload.ReportDate, payload.Shift))
        chassis_line_status = cursor.fetchall()
        # 6. Load Excel template and populate sheets
        wb = openpyxl.load_workbook(TEMPLATE_PATH)

        def get_sheet(workbook, name):
            if name not in workbook.sheetnames:
                return None
            sheet = workbook[name]
            orig_cell = sheet.cell
            def safe_cell(*args, **kwargs):
                cell = orig_cell(*args, **kwargs)
                if type(cell).__name__ == 'MergedCell':
                    for r in sheet.merged_cells.ranges:
                        if cell.coordinate in r:
                            return orig_cell(row=r.min_row, column=r.min_col)
                return cell
            sheet.cell = safe_cell
            return sheet

        # G. Update Date & Time and date formatting in summary sheets
        current_date = datetime.date.today()
        current_time = datetime.datetime.now().time()
        for sname in ["Manag Report", "PLC Work"]:
            ws = get_sheet(wb, sname)
            if ws is not None:
                ws.cell(row=4, column=2).value = current_date.strftime("%d-%m-%Y")
                ws.cell(row=5, column=2).value = current_time
                
                # Update E9 (Monthly: Short month name - YYYY, e.g. Jan 2026)
                cell_e9 = ws.cell(row=9, column=5)
                cell_e9.number_format = '@'
                cell_e9.value = report_date.strftime("%b %Y")

                # Update G9 to L9 with Yearly Week
                for i, w_header in enumerate(weekly_headers[:6]):
                    cell_w = ws.cell(row=9, column=7 + i)
                    if type(cell_w).__name__ != 'MergedCell':
                        cell_w.number_format = '@'
                        cell_w.value = w_header

                # Update E15 (Monthly: Short month name - YYYY, e.g. Jan 2026)
                cell_e15 = ws.cell(row=15, column=5)
                cell_e15.number_format = '@'
                cell_e15.value = report_date.strftime("%b %Y")
                
                # Update N9 (Daily: DD-Short month name-YYYY, e.g. 07-Jun-2026)
                cell_n9 = ws.cell(row=9, column=14)
                cell_n9.number_format = '@'
                cell_n9.value = report_date.strftime("%d-%b-%Y")

                # Update N15 (Daily: DD-Short month name-YYYY, e.g. 07-Jun-2026)
                cell_n15 = ws.cell(row=15, column=14)
                cell_n15.number_format = '@'
                cell_n15.value = report_date.strftime("%d-%b-%Y")


                # Update N23 (Daily: DD-Short month name-YYYY, e.g. 07-Jun-2026)
                cell_N23 = ws.cell(row=23, column=14)
                cell_N23.number_format = '@'
                cell_N23.value = report_date.strftime("%d-%b-%Y")

                # Update N23 (Daily: DD-Short month name-YYYY, e.g. 07-Jun-2026)
                cell_n23 = ws.cell(row=15, column=14)
                cell_n23.number_format = '@'
                cell_n23.value = report_date.strftime("%d-%b-%Y")

                # Update E45 (Daily: DD-Short month name-YYYY, e.g. 07-Jun-2026)
                cell_E45 = ws.cell(row=45, column=5)
                cell_E45.number_format = '@'
                cell_E45.value = report_date.strftime("%d-%b-%Y")
                
                # Update Q9 (MTD: DD-Short month name, e.g. 07-Jun)
                cell_q9 = ws.cell(row=9, column=17)
                cell_q9.number_format = '@'
                cell_q9.value = report_date.strftime("%d-%b")
                
                # Update U4 (Show DD YYYY)
                cell_u4 = ws.cell(row=4, column=21)
                cell_u4.number_format = '@'
                cell_u4.value = report_date.strftime("%b %Y")

                # Update U5 (Available Working Days)
                days_in_month = calendar.monthrange(report_date.year, report_date.month)[1]
                remaining_days = days_in_month - 4  # 4 सुट्ट्या वजा
                cell_u5 = ws.cell(row=5, column=21)
                cell_u5.number_format = '@'
                cell_u5.value = remaining_days

                # Update U6 (Available Working Days)
                today_day = report_date.day  # आजची तारीख
                remaining_days = max(0, days_in_month - today_day - 4)  # 4 सुट्ट्या वजा
                cell_u6 = ws.cell(row=6, column=21)
                cell_u6.number_format = '@'
                cell_u6.value = remaining_days

                # Update C15 (Show Fy YY YY - Prev Year)
                cell_c15 = ws.cell(row=15, column=3)
                cell_c15.number_format = '@'
                cell_c15.value = f"FY {(start_date.year - 1) % 100:02d}-{start_date.year % 100:02d}"

        for sname in ["Manag Report"]:
            if sname in wb.sheetnames:
                ws = wb[sname]
                #------ Saarthi Production Report TCF------
                # Update Daily Plan(Target)
                cell_N11 = ws.cell(row=11, column=14)
                cell_N11.number_format = '@'
                cell_N11.value = rows[0][1]

                # Update Daily Plan(Production)
                cell_O11 = ws.cell(row=11, column=15)
                cell_O11.number_format = '@'
                cell_O11.value = rows[0][2]

                # Update Daily Plan(Achievement)
                cell_P11 = ws.cell(row=11, column=16)
                cell_P11.number_format = '@'
                cell_P11.value = safe_int(rows[0][2]) - safe_int(rows[0][1])

                # MTD(Month To Date Actual)
                cell_Q11 = ws.cell(row=11, column=17)
                cell_Q11.number_format = '@'
                cell_Q11.value = safe_int(rows[0][17])

                # MTD(Month To Date Target)
                cell_R11 = ws.cell(row=11, column=18)
                cell_R11.number_format = '@'
                cell_R11.value = safe_int(rows[0][18])

                # MTD(Month To Date Achievement)
                cell_S11 = ws.cell(row=11, column=19)
                cell_S11.number_format = '@'
                cell_S11.value = safe_int(rows[0][18]) - safe_int(rows[0][17])

                # YTD(Year To Date Target)
                cell_T11 = ws.cell(row=11, column=20)
                cell_T11.number_format = '@'
                cell_T11.value = safe_int(rows[0][20])

                # FY(FY Year Target)
                cell_U11 = ws.cell(row=11, column=21)
                cell_U11.number_format = '@'
                cell_U11.value = safe_int(rows[0][16])

                # Week 1 (Target)
                cell_I11 = ws.cell(row=11, column=9)
                cell_I11.number_format = '@'
                cell_I11.value = safe_int(rows[0][4])

                # Week 2 (Target)
                cell_J11 = ws.cell(row=11, column=10)
                cell_J11.number_format = '@'
                cell_J11.value = safe_int(rows[0][6])

                # Week 3 (Target)
                cell_K11 = ws.cell(row=11, column=11)
                cell_K11.number_format = '@'
                cell_K11.value = safe_int(rows[0][8])

                # Week 4 (Target)
                cell_L11 = ws.cell(row=11, column=12)
                cell_L11.number_format = '@'
                cell_L11.value = safe_int(rows[0][10])

                # Week 5 (Target)
                cell_M11 = ws.cell(row=11, column=13)
                cell_M11.number_format = '@'
                cell_M11.value = safe_int(rows[0][12])

                # Monthly (Actual)
                cell_E11 = ws.cell(row=11, column=5)
                cell_E11.number_format = '@'
                cell_E11.value = safe_int(rows[0][13])

                # Monthly (Target)
                cell_F11 = ws.cell(row=11, column=6)
                cell_F11.number_format = '@'
                cell_F11.value = safe_int(rows[0][14])

                # Monthly (Achievement)
                cell_G11 = ws.cell(row=11, column=7)
                cell_G11.number_format = '@'
                cell_G11.value = safe_int(rows[0][14]) - safe_int(rows[0][13])


             #------ Micky Production Report TCF------

                # Update Daily Plan(Target)
                cell_N12 = ws.cell(row=12, column=14)
                cell_N12.number_format = '@'
                cell_N12.value = safe_int(rows1[0][1])

                # Update Daily Plan(Production)
                cell_O12 = ws.cell(row=12, column=15)
                cell_O12.number_format = '@'
                cell_O12.value = safe_int(rows1[0][2])

                # Update Daily Plan(Achievement)
                cell_P12 = ws.cell(row=12, column=16)
                cell_P12.number_format = '@'
                cell_P12.value = safe_int(rows1[0][2]) - safe_int(rows1[0][1])

                # MTD(Month To Date Actual)
                cell_Q12 = ws.cell(row=12, column=17)
                cell_Q12.number_format = '@'
                cell_Q12.value = safe_int(rows1[0][17])

                # MTD(Month To Date Target)
                cell_R12 = ws.cell(row=12, column=18)
                cell_R12.number_format = '@'
                cell_R12.value = safe_int(rows1[0][18])

                # MTD(Month To Date Achievement)
                cell_S12 = ws.cell(row=12, column=19)
                cell_S12.number_format = '@'
                cell_S12.value = safe_int(rows1[0][18]) - safe_int(rows1[0][17])

                # YTD(Year To Date Target)
                cell_T12 = ws.cell(row=12, column=20)
                cell_T12.number_format = '@'
                cell_T12.value = safe_int(rows1[0][20])

                # FY(FY Year Target)
                cell_U12 = ws.cell(row=12, column=21)
                cell_U12.number_format = '@'
                cell_U12.value = safe_int(rows1[0][16])

                # Week 1 (Target)
                cell_I12 = ws.cell(row=12, column=9)
                cell_I12.number_format = '@'
                cell_I12.value = safe_int(rows1[0][4])

                # Week 2 (Target)
                cell_J12 = ws.cell(row=12, column=10)
                cell_J12.number_format = '@'
                cell_J12.value = safe_int(rows1[0][6])

                # Week 3 (Target)
                cell_K12 = ws.cell(row=12, column=11)
                cell_K12.number_format = '@'
                cell_K12.value = safe_int(rows1[0][8])

                # Week 4 (Target)
                cell_L12 = ws.cell(row=12, column=12)
                cell_L12.number_format = '@'
                cell_L12.value = safe_int(rows1[0][10])

                # Week 5 (Target)
                cell_M12 = ws.cell(row=12, column=13)
                cell_M12.number_format = '@'
                cell_M12.value = safe_int(rows1[0][12])

                # Monthly (Actual)
                cell_E12 = ws.cell(row=12, column=5)
                cell_E12.number_format = '@'
                cell_E12.value = safe_int(rows1[0][13])

                # Monthly (Target)
                cell_F12 = ws.cell(row=12, column=6)
                cell_F12.number_format = '@'
                cell_F12.value = safe_int(rows1[0][14])

                # Monthly (Achievement)
                cell_G12 = ws.cell(row=12, column=7)
                cell_G12.number_format = '@'
                cell_G12.value = safe_int(rows1[0][14]) - safe_int(rows1[0][13])


            #------ Saarthi Production Report BIW------
                # Update Daily Plan(Production)
                cell_N17 = ws.cell(row=17, column=14)
                cell_N17.number_format = '@'
                cell_N17.value = safe_int(rows[0][2])


                # Week 1 (Production)
                cell_I17 = ws.cell(row=17, column=9)
                cell_I17.number_format = '@'
                cell_I17.value = safe_int(rows[0][4])

                # Week 2 (Production)
                cell_J17 = ws.cell(row=17, column=10)
                cell_J17.number_format = '@'
                cell_J17.value = safe_int(rows[0][6])

                # Week 3 (Production)
                cell_K17 = ws.cell(row=17, column=11)
                cell_K17.number_format = '@'
                cell_K17.value = safe_int(rows[0][8])

                # Week 4 (Production)
                cell_L17 = ws.cell(row=17, column=12)
                cell_L17.number_format = '@'
                cell_L17.value = safe_int(rows[0][10])

                # Week 5 (Production)
                cell_M17 = ws.cell(row=17, column=13)
                cell_M17.number_format = '@'
                cell_M17.value = safe_int(rows[0][12])

                # Update Daily Plan(Production)
                cell_O17 = ws.cell(row=17, column=15)
                cell_O17.number_format = '@'
                cell_O17.value = safe_int(rows1[0][2])


                # Week 1 (Production)
                cell_I21 = ws.cell(row=21, column=9)
                cell_I21.number_format = '@'
                cell_I21.value = safe_int(rows1[0][4])

                # Week 2 (Production)
                cell_J21 = ws.cell(row=21, column=10)
                cell_J21.number_format = '@'
                cell_J21.value = safe_int(rows1[0][6])

                # Week 3 (Production)
                cell_K21 = ws.cell(row=21, column=11)
                cell_K21.number_format = '@'
                cell_K21.value = safe_int(rows1[0][8])

                # Week 4 (Production)
                cell_L21 = ws.cell(row=21, column=12)
                cell_L21.number_format = '@'
                cell_L21.value = safe_int(rows1[0][10])

                # Week 5 (Production)
                cell_M21 = ws.cell(row=21, column=13)
                cell_M21.number_format = '@'
                cell_M21.value = safe_int(rows1[0][12])

                # MTD (Month To Date Actual)
                cell_G18 = ws.cell(row=18, column=7)
                cell_G18.number_format = '@'
                cell_G18.value = safe_int(rows2[0][18])

                # Monthly (Actual)
                cell_E17 = ws.cell(row=17, column=5)
                cell_E17.number_format = '@'
                cell_E17.value = safe_int(rows[0][14])

                cell_F17 = ws.cell(row=17, column=6)
                cell_F17.number_format = '@'
                cell_F17.value = safe_int(rows1[0][14])

                cell_E18 = ws.cell(row=18, column=5)
                cell_E18.number_format = '@'
                cell_E18.value = safe_int(rows2[0][14])

                cell_F17 = ws.cell(row=18, column=6)
                cell_F17.number_format = '@'
                cell_F17.value = safe_int(rows3[0][14])

                # FY Year (Actual) Chassis
                cell_C17 = ws.cell(row=17, column=3)
                cell_C17.number_format = '@'
                cell_C17.value = safe_int(rows[0][16])

                # FY Year (Actual) BIW
                cell_C18 = ws.cell(row=18, column=3)
                cell_C18.number_format = '@'
                cell_C18.value = safe_int(rows2[0][16])
                


            #------ Micky Production Report BIW----------------------
                # Update Daily Plan(Production)
                cell_N18 = ws.cell(row=18, column=14)
                cell_N18.number_format = '@'
                cell_N18.value = safe_int(rows2[0][2])

                cell_O18 = ws.cell(row=18, column=15)
                cell_O18.number_format = '@'
                cell_O18.value = safe_int(rows3[0][2])
                


                # Week 1 (Production)
                cell_I18 = ws.cell(row=18, column=9)
                cell_I18.number_format = '@'
                cell_I18.value = safe_int(rows2[0][4])

                # Week 2 (Production)
                cell_J18 = ws.cell(row=18, column=10)
                cell_J18.number_format = '@'
                cell_J18.value = safe_int(rows2[0][6])

                # Week 3 (Production)
                cell_K18 = ws.cell(row=18, column=11)
                cell_K18.number_format = '@'
                cell_K18.value = safe_int(rows2[0][8])

                # Week 4 (Production)
                cell_L18= ws.cell(row=18, column=12)
                cell_L18.number_format = '@'
                cell_L18.value = safe_int(rows2[0][10])

                # Week 5 (Production)
                cell_M18 = ws.cell(row=18, column=13)
                cell_M18.number_format = '@'
                cell_M18.value = safe_int(rows2[0][12])

                # Week 1 (Production)
                cell_I22 = ws.cell(row=22, column=9)
                cell_I22.number_format = '@'
                cell_I22.value = safe_int(rows3[0][4])

                # Week 2 (Production)
                cell_J22 = ws.cell(row=22, column=10)
                cell_J22.number_format = '@'
                cell_J22.value = safe_int(rows3[0][6])

                # Week 3 (Production)
                cell_K22 = ws.cell(row=22, column=11)
                cell_K22.number_format = '@'
                cell_K22.value = safe_int(rows3[0][8])

                # Week 4 (Production)
                cell_L22= ws.cell(row=22, column=12)
                cell_L22.number_format = '@'
                cell_L22.value = safe_int(rows3[0][10])

                # Week 5 (Production)
                cell_M22 = ws.cell(row=22, column=13)
                cell_M22.number_format = '@'
                cell_M22.value = safe_int(rows3[0][12])

                # MTD (Month To Date Actual)
                cell_H18 = ws.cell(row=18, column=8)
                cell_H18.number_format = '@'
                cell_H18.value = safe_int(rows3[0][18])

                # Monthly (Actual)
                cell_F18 = ws.cell(row=18, column=6)
                cell_F18.number_format = '@'
                cell_F18.value = safe_int(rows3[0][14])

                # FY Year (Actual) Chassis
                cell_D17 = ws.cell(row=17, column=4)
                cell_D17.number_format = '@'
                cell_D17.value = safe_int(rows1[0][16])

                # FY Year (Actual) BIW
                cell_D18 = ws.cell(row=18, column=4)
                cell_D18.number_format = '@'
                cell_D18.value = safe_int(rows3[0][16])

                # A. Populate Line Stop Daily & Monthly in 'Line Stop_Prod Loss' sheet
                ws_ls = get_sheet(wb, "Line Stop_Prod Loss")
                if ws_ls is not None:
                    
                    # 1. Daily line stops mapping
                    daily_row_mapping = {
                        "Process Call": 5,
                        "Material Call": 6,
                        "Quality Call": 7,
                        "Maintenance Call": 8,
                        "Other": 9
                    }
                    
                    for row in daily_line_stops:
                        type_of_call = row[0]
                        r = daily_row_mapping.get(type_of_call)
                        if r:
                            # SQL columns: TypeOfCallText, Chassis, Trim, Saarthi Main, Saarthi Sub, I-PUMA Main, I-PUMA Sub, Cargo Main, Cargo Sub, Chassis Line Loss Reason, Remark
                            # Map to Line Stop_Prod Loss sheet columns:
                            # Col C (3): I-PUMA Sub (row[6])
                            # Col D (4): I-PUMA Main (row[5])
                            # Col E (5): Saarthi Sub (row[4])
                            # Col F (6): Saarthi Main (row[3])
                            # Col G (7): Chassis (row[1])
                            # Col H (8): Trim (row[2])
                            val_ipuma_sub = to_mins(row[6])
                            val_ipuma_main = to_mins(row[5])
                            val_saarthi_sub = to_mins(row[4])
                            val_saarthi_main = to_mins(row[3])
                            val_chassis = to_mins(row[1])
                            val_trim = to_mins(row[2])
                            val_total = val_ipuma_sub + val_ipuma_main + val_saarthi_sub + val_saarthi_main + val_chassis + val_trim
                            
                            ws_ls.cell(row=r, column=3).value = val_ipuma_sub
                            ws_ls.cell(row=r, column=4).value = val_ipuma_main
                            ws_ls.cell(row=r, column=5).value = val_saarthi_sub
                            ws_ls.cell(row=r, column=6).value = val_saarthi_main
                            ws_ls.cell(row=r, column=7).value = val_chassis
                            ws_ls.cell(row=r, column=8).value = val_trim
                            ws_ls.cell(row=r, column=9).value = val_total
                            
                            # Chassis Line Loss Reason (Col N / 14)
                            ws_ls.cell(row=r, column=14).value = row[9]
                            
                    # 2. Monthly line stops mapping
                    # First aggregate raw records
                    def get_station_group(station_no):
                        if station_no is None:
                            return None
                        try:
                            s = int(station_no)
                        except ValueError:
                            return None
                        if 31 <= s <= 36:
                            return "Chassis"
                        elif 25 <= s <= 30:
                            return "Trim"
                        elif 18 <= s <= 24:
                            return "Saarthi Main"
                        elif 14 <= s <= 17:
                            return "Saarthi Sub"
                        elif 5 <= s <= 13:
                            return "I-PUMA Main"
                        elif 1 <= s <= 4:
                            return "I-PUMA Sub"
                        return None

                    monthly_data = {
                        cat: {
                            "Chassis": 0.0,
                            "Trim": 0.0,
                            "Saarthi Main": 0.0,
                            "Saarthi Sub": 0.0,
                            "I-PUMA Main": 0.0,
                            "I-PUMA Sub": 0.0,
                            "reasons": {
                                "Chassis": {},
                                "Trim": {},
                                "Saarthi Main": {},
                                "Saarthi Sub": {},
                                "I-PUMA Main": {},
                                "I-PUMA Sub": {}
                            }
                        } for cat in ["Process Call", "Material Call", "Quality Call", "Maintenance Call", "Other"]
                    }

                    for row in monthly_line_stops:
                        # SQL monthly row columns: DT, LineStopTime, TypeOfCall, TypeOfCallText, ReasonCode, ReasonText, StationNo
                        loss_time = float(row[1]) if row[1] is not None else 0.0
                        type_of_call = row[3]
                        reason = row[5]
                        station_no = row[6]
                        group = get_station_group(station_no)
                        
                        if type_of_call not in monthly_data:
                            type_of_call = "Other"
                        
                        if group is not None:
                            monthly_data[type_of_call][group] += loss_time
                            if reason is not None and str(reason).strip() not in ('', '0', 'null'):
                                r_dict = monthly_data[type_of_call]["reasons"][group]
                                r_dict[reason] = r_dict.get(reason, 0.0) + loss_time

                    monthly_row_mapping = {
                        "Process Call": 24,
                        "Material Call": 25,
                        "Quality Call": 26,
                        "Maintenance Call": 27,
                        "Other": 28
                    }

                    for cat, r in monthly_row_mapping.items():
                        val_ipuma_sub = to_mins(monthly_data[cat]["I-PUMA Sub"])
                        val_ipuma_main = to_mins(monthly_data[cat]["I-PUMA Main"])
                        val_saarthi_sub = to_mins(monthly_data[cat]["Saarthi Sub"])
                        val_saarthi_main = to_mins(monthly_data[cat]["Saarthi Main"])
                        val_chassis = to_mins(monthly_data[cat]["Chassis"])
                        val_trim = to_mins(monthly_data[cat]["Trim"])
                        val_total = val_ipuma_sub + val_ipuma_main + val_saarthi_sub + val_saarthi_main + val_chassis + val_trim
                        
                        ws_ls.cell(row=r, column=3).value = val_ipuma_sub
                        ws_ls.cell(row=r, column=4).value = val_ipuma_main
                        ws_ls.cell(row=r, column=5).value = val_saarthi_sub
                        ws_ls.cell(row=r, column=6).value = val_saarthi_main
                        ws_ls.cell(row=r, column=7).value = val_chassis
                        ws_ls.cell(row=r, column=8).value = val_trim
                        ws_ls.cell(row=r, column=9).value = val_total

                        # Write top reasons
                        # Map station groups to column indices in sheet:
                        # I-PUMA Sub -> J (10)
                        # I-PUMA Main -> K (11)
                        # Saarthi Sub -> L (12)
                        # Saarthi Main -> M (13)
                        # Chassis -> N (14)
                        # Trim -> O (15)
                        top_reason_col_mapping = {
                            "I-PUMA Sub": 10,
                            "I-PUMA Main": 11,
                            "Saarthi Sub": 12,
                            "Saarthi Main": 13,
                            "Chassis": 14,
                            "Trim": 15
                        }
                        for group, col_idx in top_reason_col_mapping.items():
                            reasons_dict = monthly_data[cat]["reasons"][group]
                            if reasons_dict:
                                top_reason = max(reasons_dict, key=reasons_dict.get)
                                ws_ls.cell(row=r, column=col_idx).value = top_reason
                            else:
                                ws_ls.cell(row=r, column=col_idx).value = None

                # B. Populate Cargo Main/Sub, Chassis Line Loss Reason and Remarks in 'Manag Report' sheet
                ws_man = get_sheet(wb, "Manag Report")
                if ws_man is not None:
                    
                    # Map Cargo Main/Sub and Chassis details for Daily section in Manag Report (rows 48-52)
                    man_row_mapping = {
                        "Process Call": 48,
                        "Material Call": 49,
                        "Quality Call": 50,
                        "Maintenance Call": 51,
                        "Other": 52
                    }
                    
                    for row in daily_line_stops:
                        type_of_call = row[0]
                        r = man_row_mapping.get(type_of_call)
                        if r:
                            # Col 11 (K): Cargo Main (row[7])
                            ws_man.cell(row=r, column=11).value = to_mins(row[7])
                            # Col 13 (M): Cargo Sub (row[8])
                            ws_man.cell(row=r, column=13).value = to_mins(row[8])
                            # Col 15 (O): Chassis Line Loss Reason (row[9])
                            ws_man.cell(row=r, column=15).value = row[9]
                            # Col 19 (S): Remark (row[10])
                            ws_man.cell(row=r, column=19).value = row[10]

                    # C. Populate Chassis Line Status & Chassis Production Daily in 'Manag Report' sheet
                    if chassis_line_status:
                        # 1. Update row 34: Chassis Production Daily
                        first_row = chassis_line_status[0]
                        # Col D (4): Shift
                        ws_man.cell(row=34, column=4).value = first_row[9]
                        # Col E (5): Shift Start Time
                        ws_man.cell(row=34, column=5).value = to_hours(first_row[10])
                        # Col F (6): Shift End Time
                        ws_man.cell(row=34, column=6).value = to_hours(first_row[11])
                        # Col G (7): Shift Time (Min)
                        ws_man.cell(row=34, column=7).value = safe_int(first_row[14])
                        # Col H (8): Plan Down Time (Min) (DownTime)
                        ws_man.cell(row=34, column=8).value = safe_int(first_row[17])
                        # Col I (9): Production Count
                        ws_man.cell(row=34, column=9).value = safe_int(first_row[12])
                        # Col J (10): OEE from database, formatted as 4 decimal places
                        oee_val = first_row[19]
                        cell_J34 = ws_man.cell(row=34, column=10)
                        cell_J34.number_format = '0.0000'
                        cell_J34.value = round(float(oee_val), 4) if oee_val is not None else 0.0

                        # 2. Update rows 38-43: Chassis Line Status table
                        station_row_mapping = {
                            "CH-10": 38,
                            "CH-20": 39,
                            "CH-30": 40,
                            "CH-40": 41,
                            "CH-50": 42,
                            "CH-60": 43
                        }
                        for row in chassis_line_status:
                            station_num = row[0]
                            r = station_row_mapping.get(station_num)
                            if r:
                                # Col B (2): Production Loss (Min) -> TotalLoss
                                ws_man.cell(row=r, column=2).value = to_mins(row[6])
                                # Col C (3): Call Loss Process (Min)
                                ws_man.cell(row=r, column=3).value = to_mins(row[1])
                                # Col D (4): Call Loss Material (Min)
                                ws_man.cell(row=r, column=4).value = to_mins(row[2])
                                # Col E (5): Call Loss Quality (Min)
                                ws_man.cell(row=r, column=5).value = to_mins(row[3])
                                # Col F (6): Call Loss Maint (Min)
                                ws_man.cell(row=r, column=6).value = to_mins(row[4])
                                # Col G (7): Call Loss Other (Min)
                                ws_man.cell(row=r, column=7).value = to_mins(row[5])
                                # Col H (8): Call Loss Remark
                                ws_man.cell(row=r, column=8).value = row[8]

            
        # 7. Save and return workbook
        wb.save(output_path)

        return {
            "status": "success",
            "filepath": str(output_path),
            "filename": output_path.name
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/open-file")
def open_file_endpoint(payload: OpenFileRequest):
    try:
        path = os.path.abspath(payload.filepath)
        if not os.path.exists(path):
            raise HTTPException(status_code=404, detail="File not found")
        
        if os.name == "nt":
            os.startfile(path)
        else:
            opener = "open" if sys.platform == "darwin" else "xdg-open"
            subprocess.call([opener, path])
            
        return {"status": "success", "message": "File opened successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/open-folder")
def open_folder_endpoint(payload: OpenFileRequest):
    try:
        path = os.path.abspath(payload.filepath)
        if not os.path.exists(path):
            raise HTTPException(status_code=404, detail="File not found")
        
        if os.name == "nt":
            # Highlight the file in Explorer
            subprocess.Popen(f'explorer /select,"{path}"')

        return {"status": "success", "message": "Folder opened successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))