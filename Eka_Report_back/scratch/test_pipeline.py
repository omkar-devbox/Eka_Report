import openpyxl
import datetime

# 1. Load template
wb = openpyxl.load_workbook("ProductionReport_R3.xlsx")

# Check images before any operations
print("Images in template before any operations:")
for sname in wb.sheetnames:
    print(f"  {sname}: {len(getattr(wb[sname], '_images', []))}")

# 2. Replicate get_sheet override
def get_sheet(workbook, name):
    if name not in workbook.sheetnames:
        return None
    sheet = workbook[name]
    orig_cell = sheet.cell
    def safe_cell(*args, **kwargs):
        cell = orig_cell(*args, **kwargs)
        if type(cell).__name__ == 'MergedCell':
            for r in sheet.merged_cells.ranges:
                if cell.coordinate in r:
                    return orig_cell(row=r.min_row, column=r.min_col)
        return cell
    sheet.cell = safe_cell
    return sheet

# 3. Modify cells
current_date = datetime.date.today()
current_time = datetime.datetime.now().time()
report_date = datetime.date(2026, 6, 13)

for sname in ["Manag Report", "Prod Report"]:
    ws = get_sheet(wb, sname)
    if ws is not None:
        cell_e2 = ws.cell(row=2, column=5)
        cell_e2.value = current_date.strftime("%d-%m-%Y")
        
        # Test modifying something on ws
        ws.cell(row=10, column=4).value = 100

# Check images before saving
print("Images in workbook before saving:")
for sname in wb.sheetnames:
    print(f"  {sname}: {len(getattr(wb[sname], '_images', []))}")

# 4. Save workbook
wb.save("scratch/test_pipeline_out.xlsx")

# 5. Load and check images after saving
print("Images in saved workbook:")
wb_saved = openpyxl.load_workbook("scratch/test_pipeline_out.xlsx")
for sname in wb_saved.sheetnames:
    print(f"  {sname}: {len(getattr(wb_saved[sname], '_images', []))}")
