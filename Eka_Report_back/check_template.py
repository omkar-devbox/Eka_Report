import openpyxl

wb = openpyxl.load_workbook('ProductionReport_R3.xlsx')
ws = wb['Prod Report']

print("=== Prod Report Sheet non-empty cells ===")
for r in range(1, 150):
    row_str = []
    has_content = False
    for c in range(1, 25):
        val = ws.cell(row=r, column=c).value
        if val is not None:
            has_content = True
            row_str.append(f"{openpyxl.utils.get_column_letter(c)}{r}:{val}")
    if has_content:
        print(f"Row {r}: " + " | ".join(row_str))
