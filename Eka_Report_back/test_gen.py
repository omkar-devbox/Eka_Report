import sys
import os
import datetime

# Add current directory to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app.core.database import get_db_connection
from app.schemas.mgmt_prod_report import MgmtProdReportRequest
from app.api.MgmtProdReport.mgmt_prod_report import generate_mgmt_production_report

def main():
    print("Initializing test run...")
    # Mock payload for report generation
    # Today is June 9, 2026. The FY is April 1, 2026 to March 31, 2027.
    payload = MgmtProdReportRequest(
        ReportDate="2026-06-09",
        StartDate="2026-04-01",
        LastDate="2027-03-31",
        Shift="All"
    )
    
    gen = get_db_connection()
    conn = next(gen)
    
    try:
        print("Generating report...")
        result = generate_mgmt_production_report(payload, conn)
        print("Success:", result)
        
        # Open and inspect the generated Excel sheet
        import openpyxl
        fpath = result["filepath"]
        wb = openpyxl.load_workbook(fpath)
        
        print("\n--- Inspecting SQL Work ---")
        ws_sql = wb["SQL Work"]
        print("C15:", ws_sql.cell(row=15, column=3).value)
        print("C17 (Saarthi Chassis):", ws_sql.cell(row=17, column=3).value)
        print("C18 (Saarthi BIW):", ws_sql.cell(row=18, column=3).value)
        print("D17 (Micky Chassis):", ws_sql.cell(row=17, column=4).value)
        print("D18 (Micky BIW):", ws_sql.cell(row=18, column=4).value)
        
        print("\n--- Inspecting PLC Work ---")
        ws_plc = wb["PLC Work"]
        print("C15:", ws_plc.cell(row=15, column=3).value)
        print("C17:", ws_plc.cell(row=17, column=3).value)
        print("C18:", ws_plc.cell(row=18, column=3).value)
        print("D17:", ws_plc.cell(row=17, column=4).value)
        print("D18:", ws_plc.cell(row=18, column=4).value)

        print("\n--- Inspecting PrevFY ---")
        ws_prev = wb["PrevFY"]
        print("Row 3 (M_TCF):", [ws_prev.cell(row=3, column=c).value for c in range(1, 10)])
        print("Row 12 (S_TCF):", [ws_prev.cell(row=12, column=c).value for c in range(1, 10)])
        print("Row 22 (M_BIW):", [ws_prev.cell(row=22, column=c).value for c in range(1, 10)])
        print("Row 32 (S_BIW):", [ws_prev.cell(row=32, column=c).value for c in range(1, 10)])
        
    except Exception as err:
        print("Error occurred:", err)
    finally:
        try:
            next(gen)
        except StopIteration:
            pass

if __name__ == "__main__":
    main()
